#!/usr/bin/env python3
"""Extract Scott's dealership reports into normalized JSON for the dashboard."""
import json, csv, re
import openpyxl

SRC = "/Users/carlos/Desktop/Scott"
OUT = __import__("os").path.join(__import__("os").path.dirname(__import__("os").path.abspath(__file__)), "data.json")


def read_report(fname):
    """Return (rows:list[dict], filters:dict) from a VinSolutions export."""
    wb = openpyxl.load_workbook(f"{SRC}/{fname}", read_only=True, data_only=True)
    ws = wb["Report"]
    raw = [r for r in ws.iter_rows(values_only=True) if any(v is not None and str(v).strip() != "" for v in r)]
    header = [str(h).strip() for h in raw[0]]
    rows = []
    for r in raw[1:]:
        d = {}
        for h, v in zip(header, r):
            d[h] = v
        rows.append(d)
    filters = {}
    wsf = wb["Filters"]
    for r in wsf.iter_rows(values_only=True):
        if r and r[0] and str(r[0]).strip() and str(r[0]).strip() != "Filter Name":
            filters[str(r[0]).strip()] = str(r[2]).strip() if len(r) > 2 and r[2] is not None else ""
    wb.close()
    return rows, filters


def num(v):
    if v is None or v == "":
        return 0
    f = float(v)
    return int(f) if f == int(f) else round(f, 4)


def kpi_metrics(row):
    """Normalize a KPI row's metric columns (present columns vary by report)."""
    m = {
        "goodLeads": num(row.get("Good Leads")),
        "sold": num(row.get("Sold in Time Frame")),
        "contactPct": num(row.get("Internet Actual Contact %")),
        "apptSetOfContactedPct": num(row.get("Appts Set of Contacted %")),
        "apptsShown": num(row.get("Appts Shown")),
    }
    if "Appts Set %" in row:
        m["apptSetPct"] = num(row.get("Appts Set %"))
    return m


def parse_kpi_leadtype(fname):
    """Parse a KPI report summarized by Lead Type / Inventory Type (dealer col optional).
    Returns {"total": metrics, "byLeadType": [{leadType, metrics, byInventory:[...]}]}"""
    rows, filters = read_report(fname)
    total = None
    by_lt = []
    cur = None
    for r in rows:
        lt = (r.get("Lead Type") or "").strip()
        inv = (r.get("Inventory Type") or "").strip()
        dealer = (r.get("Dealer") or "").strip()
        first_col = dealer or lt
        if first_col == "TOTAL" or (dealer == "TOTAL"):
            total = kpi_metrics(r)
        elif dealer and not lt and not inv:
            # dealer-level total row (Dealer summary reports)
            total = total or kpi_metrics(r)
        elif lt and lt != "TOTAL" and not inv:
            cur = {"leadType": lt, "metrics": kpi_metrics(r), "byInventory": []}
            by_lt.append(cur)
        elif lt and inv and cur is not None and cur["leadType"] == lt:
            cur["byInventory"].append({"inventoryType": inv, "metrics": kpi_metrics(r)})
    return {"total": total, "byLeadType": by_lt}, filters


def parse_kpi_make(fname):
    """Parse the Vern Eide KPI report: Lead Type / Inventory Type / Vehicle Make."""
    rows, filters = read_report(fname)
    total = None
    by_lt = []
    cur_lt = None
    cur_inv = None
    for r in rows:
        lt = (r.get("Lead Type") or "").strip()
        inv = (r.get("Inventory Type") or "").strip()
        make = (r.get("Vehicle Make") or "").strip()
        if lt == "TOTAL":
            total = kpi_metrics(r)
        elif lt and not inv and not make:
            cur_lt = {"leadType": lt, "metrics": kpi_metrics(r), "byInventory": []}
            by_lt.append(cur_lt)
        elif lt and inv and not make:
            cur_inv = {"inventoryType": inv, "metrics": kpi_metrics(r), "byMake": []}
            cur_lt["byInventory"].append(cur_inv)
        elif lt and inv and make:
            cur_inv["byMake"].append({"make": make, "metrics": kpi_metrics(r)})
    return {"total": total, "byLeadType": by_lt}, filters


HOUSE_ACCOUNT = re.compile(r"\b(team|house)\b", re.IGNORECASE)


def parse_salesteam(fname):
    rows, filters = read_report(fname)
    team, totals = [], None
    for r in rows:
        u = (r.get("User") or "").strip()
        if not u:
            continue
        # skip CRM house accounts ("Vern Eide Team", "House Deal", ...) — not people
        if u != "TOTAL" and HOUSE_ACCOUNT.search(u):
            continue
        rec = {
            "name": u,
            "goodLeads": num(r.get("Good Leads")),
            "sold": num(r.get("Sold in Time Frame")),
            "apptsScheduled": num(r.get("Appts Scheduled")),
            "apptsShown": num(r.get("Appts Shown")),
            "shownPct": num(r.get("Appts Shown %")),
            "calls": num(r.get("Calls Out")),
            "emails": num(r.get("Emails Out")),
            "texts": num(r.get("Texts Out")),
        }
        if u == "TOTAL":
            totals = rec
        else:
            team.append(rec)
    return team, totals, filters


def parse_matador():
    out = []
    with open(f"{SRC}/Matador MTD Stats Vern Eide Honda.csv", newline="", encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            out.append({
                "name": r["User"].strip(),
                "role": r["User Role"].strip(),
                "lastActivity": r["Last Activity"].strip()[:10],
                "apptsCreated": int(r["Appointments Created"]),
                "videosSent": int(r["Videos Sent"]),
                "messagesSent": int(r["Sent Messages"]),
                "reviewInvites": int(r["Review Invites Sent"]),
                "clientsMessaged": int(r["Clients Messaged"]),
                "assignedClients": int(r["Assigned Clients"]),
            })
    return out


def norm_name(n):
    n = re.sub(r"\s+", " ", n.strip().lower())
    return n

# manual aliases: Matador name -> CRM name
ALIASES = {"zach schroeder": "zachary schroeder", "scott mcdonald": "scott mcdonald"}

# ---------------- build ----------------
vern_kpi, vern_kpi_f = parse_kpi_make("Vern_Eide_Honda_KPI_Stats_2026-06-22 (1).xlsx")
vern_team, vern_totals, vern_team_f = parse_salesteam("Vern_Eide_Honda_MTD_Sales_Stats_2026-06-22.xlsx")
matador = parse_matador()

n777_cur, f777_cur = parse_kpi_leadtype("Report-8438.xlsx")
n777_pri, f777_pri = parse_kpi_leadtype("Report-5113.xlsx")
n777_team, n777_totals, f777_team = parse_salesteam("Report-3321.xlsx")

arm_cur, farm_cur = parse_kpi_leadtype("Report-7119.xlsx")
arm_pri, farm_pri = parse_kpi_leadtype("Report-984.xlsx")

# join Matador onto Vern Eide CRM team
mat_by_name = {}
for m in matador:
    key = ALIASES.get(norm_name(m["name"]), norm_name(m["name"]))
    mat_by_name[key] = m
matched = set()
for rep in vern_team:
    m = mat_by_name.get(norm_name(rep["name"]))
    if m:
        rep["matador"] = {k: m[k] for k in ("messagesSent", "videosSent", "clientsMessaged", "assignedClients", "apptsCreated", "lastActivity")}
        matched.add(norm_name(m["name"]))
matador_other = [m for m in matador if norm_name(m["name"]) not in matched and ALIASES.get(norm_name(m["name"]), norm_name(m["name"])) not in matched]

data = {
    "generatedAt": "2026-07-03",
    "stores": [
        {
            "id": "vern-eide-honda", "name": "Vern Eide Honda", "demo": False,
            "crm": "VinSolutions", "tools": ["Matador"],
            "location": "Sioux Falls, SD",
            "currentPeriod": {"label": "Jun 1–22, 2026 (MTD)", "exported": "Jun 22, 2026"},
            "priorPeriod": None,
            "kpiCurrent": vern_kpi, "kpiPrior": None,
            "salesTeam": vern_team, "salesTeamTotals": vern_totals,
            "salesTeamPeriod": "Jun 1–22, 2026 (MTD)",
            "matadorOther": matador_other,
            "notes": "Source: VinSolutions Enterprise Performance exports + Matador Users report, exported Jun 22, 2026. No prior-month baseline file was provided for this store.",
        },
        {
            "id": "777-nissan", "name": "777 Nissan", "demo": False,
            "crm": "VinSolutions", "tools": [],
            "location": "—",
            "currentPeriod": {"label": "Jul 1–2, 2026 (MTD)", "exported": "Jul 2, 2026"},
            "priorPeriod": {"label": "Jun 1–2, 2026 (same days)"},
            "kpiCurrent": n777_cur, "kpiPrior": n777_pri,
            "salesTeam": n777_team, "salesTeamTotals": n777_totals,
            "salesTeamPeriod": "Jul 1–2, 2026 (MTD)",
            "matadorOther": [],
            "notes": "MTD KPI Comparison: current month-to-date (Jul 1–2) vs the same days of the previous month (Jun 1–2). Exported Jul 2, 2026.",
        },
        {
            "id": "armstrong-subaru", "name": "Armstrong Subaru", "demo": False,
            "crm": "VinSolutions", "tools": [],
            "location": "—",
            "currentPeriod": {"label": "Jul 1–2, 2026 (MTD)", "exported": "Jul 2, 2026"},
            "priorPeriod": {"label": "Jun 1–2, 2026 (same days)"},
            "kpiCurrent": arm_cur, "kpiPrior": arm_pri,
            "salesTeam": None, "salesTeamTotals": None, "salesTeamPeriod": None,
            "matadorOther": [],
            "notes": "MTD KPI Comparison: Jul 1–2 vs Jun 1–2. No salesperson activity report was provided for this store.",
        },
    ],
    "integrations": [
        {"name": "VinSolutions", "type": "CRM", "coverage": "Majority of stores", "api": "requested", "scheduledEmail": True,
         "note": "Existing reports can be scheduled to be emailed daily. API access has been requested."},
        {"name": "Tekion", "type": "CRM", "coverage": "Some stores", "api": "requested", "scheduledEmail": True,
         "note": "Has an API — access inquiry in progress. Reports can be scheduled via email."},
        {"name": "DriveCentric", "type": "CRM", "coverage": "Some stores", "api": "unknown", "scheduledEmail": True,
         "note": "API availability unconfirmed. Reports can be scheduled."},
        {"name": "Momentum", "type": "CRM", "coverage": "Some stores", "api": "requested", "scheduledEmail": True,
         "note": "Has an API — asking about access. Reports can be scheduled."},
        {"name": "Matador", "type": "AI messaging", "coverage": "~50% of stores", "api": "unknown", "scheduledEmail": False,
         "note": "Activity report is downloaded manually from the dashboard Users tab. Scheduled email not available today."},
        {"name": "Covideo", "type": "Video outreach", "coverage": "Some stores", "api": "unknown", "scheduledEmail": True,
         "note": "Activity reporting can be scheduled via email."},
    ],
}

# ---- demo stores (clearly flagged) to preview the multi-CRM rollout ----
def demo_kpi(leads, sold, contact, apptset, shown, mix):
    return {
        "total": {"goodLeads": leads, "sold": sold, "contactPct": contact, "apptSetOfContactedPct": apptset, "apptsShown": shown},
        "byLeadType": [
            {"leadType": lt, "metrics": {"goodLeads": l, "sold": s, "contactPct": c, "apptSetOfContactedPct": a, "apptsShown": sh}, "byInventory": []}
            for (lt, l, s, c, a, sh) in mix
        ],
    }

data["stores"] += [
    {
        "id": "demo-summit-toyota", "name": "Summit Toyota", "demo": True,
        "crm": "Tekion", "tools": ["Covideo"], "location": "Demo data",
        "currentPeriod": {"label": "Jun 1–22, 2026 (MTD)", "exported": "—"},
        "priorPeriod": {"label": "May 1–22, 2026 (same days)"},
        "kpiCurrent": demo_kpi(412, 118, 0.64, 0.41, 96, [
            ("Internet", 180, 31, 0.64, 0.38, 34), ("Phone", 84, 27, 0, 0.55, 30),
            ("Walk-in", 118, 51, 0, 0.24, 24), ("PreviousCustomer", 30, 9, 0, 0.27, 8)]),
        "kpiPrior": demo_kpi(390, 104, 0.58, 0.37, 84, [
            ("Internet", 171, 26, 0.58, 0.35, 30), ("Phone", 80, 24, 0, 0.51, 26),
            ("Walk-in", 111, 46, 0, 0.22, 21), ("PreviousCustomer", 28, 8, 0, 0.25, 7)]),
        "salesTeam": None, "salesTeamTotals": None, "salesTeamPeriod": None, "matadorOther": [],
        "notes": "DEMO store — illustrative data only, shows how a Tekion-based store plugs into the same dashboard.",
    },
    {
        "id": "demo-riverbend-cdjr", "name": "Riverbend CDJR", "demo": True,
        "crm": "DriveCentric", "tools": ["Matador"], "location": "Demo data",
        "currentPeriod": {"label": "Jun 1–22, 2026 (MTD)", "exported": "—"},
        "priorPeriod": {"label": "May 1–22, 2026 (same days)"},
        "kpiCurrent": demo_kpi(298, 61, 0.51, 0.29, 48, [
            ("Internet", 152, 18, 0.51, 0.27, 20), ("Phone", 55, 14, 0, 0.42, 14),
            ("Walk-in", 74, 24, 0, 0.2, 11), ("Referral", 17, 5, 0, 0.3, 3)]),
        "kpiPrior": demo_kpi(340, 79, 0.57, 0.34, 61, [
            ("Internet", 170, 24, 0.57, 0.31, 26), ("Phone", 62, 18, 0, 0.47, 18),
            ("Walk-in", 88, 31, 0, 0.24, 14), ("Referral", 20, 6, 0, 0.33, 3)]),
        "salesTeam": None, "salesTeamTotals": None, "salesTeamPeriod": None, "matadorOther": [],
        "notes": "DEMO store — illustrative data only (DriveCentric CRM + Matador).",
    },
    {
        "id": "demo-lakeside-ford", "name": "Lakeside Ford", "demo": True,
        "crm": "Momentum", "tools": ["Matador", "Covideo"], "location": "Demo data",
        "currentPeriod": {"label": "Jun 1–22, 2026 (MTD)", "exported": "—"},
        "priorPeriod": {"label": "May 1–22, 2026 (same days)"},
        "kpiCurrent": demo_kpi(355, 97, 0.68, 0.44, 88, [
            ("Internet", 168, 29, 0.68, 0.41, 36), ("Phone", 71, 25, 0, 0.6, 26),
            ("Walk-in", 92, 36, 0, 0.28, 20), ("PreviousCustomer", 24, 7, 0, 0.29, 6)]),
        "kpiPrior": demo_kpi(348, 92, 0.66, 0.45, 86, [
            ("Internet", 165, 28, 0.66, 0.42, 35), ("Phone", 70, 24, 0, 0.6, 25),
            ("Walk-in", 90, 34, 0, 0.28, 20), ("PreviousCustomer", 23, 6, 0, 0.3, 6)]),
        "salesTeam": None, "salesTeamTotals": None, "salesTeamPeriod": None, "matadorOther": [],
        "notes": "DEMO store — illustrative data only (Momentum CRM + Matador + Covideo).",
    },
]

with open(OUT, "w") as f:
    json.dump(data, f, indent=1)
print("wrote", OUT)
# quick sanity summary
for s in data["stores"]:
    kc = s["kpiCurrent"]["total"]
    print(f'{s["name"]:22} demo={s["demo"]} leads={kc["goodLeads"]} sold={kc["sold"]} shown={kc["apptsShown"]} team={len(s["salesTeam"]) if s["salesTeam"] else 0} lts={[b["leadType"] for b in s["kpiCurrent"]["byLeadType"]]}')
