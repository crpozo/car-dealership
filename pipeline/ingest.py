#!/usr/bin/env python3
"""Ingest VinSolutions report exports into pipeline/data.json.

Sources it walks, recursively, in any mix:
  *.zip    a Google Takeout archive (extracted to a cache dir, then re-walked)
  *.mbox   a Takeout mail export — every .xlsx attachment is pulled out
  *.xlsx   a loose report export
  *.csv    a Matador "Users" activity export

Every workbook is classified from its **Filters** sheet, never from the filename or
the mail subject: Dealers / Date Range / Date Range Begin / Date Range End / Run Date /
Summary Level 1. The same report is often sent twice, so snapshots are de-duplicated on
(storeId, kind, period, begin, end, runDate).

Usage:
    python3 pipeline/ingest.py [SRC_DIR ...]     # default: ~/Desktop/Scott
    python3 pipeline/build.py                    # then: data.json -> assets/data.js
"""
import csv
import json
import os
import re
import shutil
import sys
import zipfile
from collections import defaultdict
from datetime import datetime

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "data.json")
DEFAULT_SRC = os.path.expanduser("~/Desktop/Scott")
CACHE = os.path.join(HERE, ".cache")

# Stores kept out of the dashboard entirely, by slug. Vern Eide Honda is not on the
# scheduled-report list, so its only export is a one-off from Jun 22 2026 — a stale
# store on the roster is worse than no store. Delete the entry (and re-run) if it
# ever starts sending reports. Excluded stores are reported, never dropped silently.
EXCLUDE_STORES = {"vern-eide-honda"}

MONTHS = "Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split()
HOUSE_ACCOUNT = re.compile(r"\b(team|house)\b", re.IGNORECASE)
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# --------------------------------------------------------------------------- #
# small helpers
# --------------------------------------------------------------------------- #

def slug(name):
    s = re.sub(r"[^a-z0-9]+", "-", (name or "").strip().lower())
    return s.strip("-")


def parse_dt(value):
    """'Jul  2 2026  8:10AM' / 'Jun  1 2026 12:00AM' -> date string 'YYYY-MM-DD'."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    parts = re.sub(r"\s+", " ", str(value).strip()).split(" ")
    if len(parts) < 3 or parts[0][:3] not in MONTHS:
        return None
    try:
        return datetime(int(parts[2]), MONTHS.index(parts[0][:3]) + 1, int(parts[1])).strftime("%Y-%m-%d")
    except ValueError:
        return None


def num(value):
    if value is None or value == "":
        return 0
    try:
        f = float(value)
    except (TypeError, ValueError):
        return 0
    return int(f) if f == int(f) else round(f, 6)


def cell(row, idx, col):
    if col not in idx:
        return ""
    v = row[idx[col]]
    return "" if v is None else str(v).strip()


# --------------------------------------------------------------------------- #
# source discovery: zip / mbox / xlsx / csv
# --------------------------------------------------------------------------- #

def unpack_zips(paths, log):
    """Extract every zip into CACHE and return the extra roots to walk."""
    roots = []
    for p in paths:
        target = os.path.join(CACHE, "zip", slug(os.path.basename(p))[:60])
        try:
            shutil.rmtree(target, ignore_errors=True)
            os.makedirs(target, exist_ok=True)
            with zipfile.ZipFile(p) as zf:
                zf.extractall(target)
            roots.append(target)
            log.append("unzipped %s" % os.path.basename(p))
        except Exception as exc:  # noqa: BLE001 - a bad archive must not abort the run
            log.append("SKIP zip %s: %s" % (os.path.basename(p), exc))
    return roots


def extract_mbox(path, log):
    """Pull every .xlsx attachment out of an mbox. Returns list of file paths."""
    import mailbox

    out_dir = os.path.join(CACHE, "mbox", slug(os.path.basename(path))[:60])
    shutil.rmtree(out_dir, ignore_errors=True)
    os.makedirs(out_dir, exist_ok=True)
    found, seen = [], set()
    try:
        box = mailbox.mbox(path)
    except Exception as exc:  # noqa: BLE001
        log.append("SKIP mbox %s: %s" % (os.path.basename(path), exc))
        return []
    for msg in box:
        for part in msg.walk():
            fname = part.get_filename()
            if not fname or not fname.lower().endswith(".xlsx"):
                continue
            try:
                payload = part.get_payload(decode=True)
            except Exception:  # noqa: BLE001
                continue
            if not payload:
                continue
            name, i = fname, 1
            while name in seen:
                name = "%s__%d.xlsx" % (fname[:-5], i)
                i += 1
            seen.add(name)
            dest = os.path.join(out_dir, name)
            with open(dest, "wb") as fh:
                fh.write(payload)
            found.append(dest)
    log.append("mbox %s -> %d xlsx attachments" % (os.path.basename(path), len(found)))
    return found


def discover(src_dirs, log):
    """Walk the sources and return (xlsx_paths, matador_csv_paths)."""
    xlsx, csvs, mboxes, zips = [], [], [], []

    def walk(roots):
        for root in roots:
            if os.path.isfile(root):
                classify(root)
                continue
            for dirpath, dirnames, filenames in os.walk(root):
                dirnames[:] = [d for d in dirnames if d != os.path.basename(CACHE)]
                for fn in filenames:
                    classify(os.path.join(dirpath, fn))

    def classify(path):
        low = path.lower()
        base = os.path.basename(path)
        if base.startswith("~$") or base.startswith("."):
            return
        if low.endswith(".xlsx"):
            xlsx.append(path)
        elif low.endswith(".mbox"):
            mboxes.append(path)
        elif low.endswith(".zip"):
            zips.append(path)
        elif low.endswith(".csv") and "matador" in base.lower():
            csvs.append(path)

    walk(src_dirs)
    if zips:
        walk(unpack_zips(zips, log))
    for mb in mboxes:
        xlsx.extend(extract_mbox(mb, log))
    return xlsx, csvs


# --------------------------------------------------------------------------- #
# workbook parsing
# --------------------------------------------------------------------------- #

def read_filters(wb):
    out = {}
    if "Filters" not in wb.sheetnames:
        return out
    for row in wb["Filters"].iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = re.sub(r"\s+", " ", str(row[0])).strip()
        if not key or key == "Filter Name":
            continue
        out[key] = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
    return out


def read_report(wb):
    rows = [r for r in wb["Report"].iter_rows(values_only=True)
            if any(v is not None and str(v).strip() != "" for v in r)]
    if not rows:
        return [], {}
    header = [re.sub(r"\s+", " ", str(h)).strip() if h is not None else "" for h in rows[0]]
    return rows[1:], {h: i for i, h in enumerate(header) if h}


def metrics_from(row, idx):
    """Normalize one report row into the metrics bag the dashboard consumes.

    "Internet Actual Contact %" is an internet-only measure: VinSolutions writes a
    literal 0 for Phone / Walk-in / Referral / PreviousCustomer rows, and copies the
    *internet* rate onto the store TOTAL row. So the derived contacted count is only
    meaningful where the rate is non-zero, and the TOTAL row's count has to be based
    on internet good leads, not on all good leads — see reconcile_total() below.
    """
    good = num(row[idx["Good Leads"]]) if "Good Leads" in idx else 0
    contact_pct = num(row[idx["Internet Actual Contact %"]]) if "Internet Actual Contact %" in idx else 0
    set_of_contacted = num(row[idx["Appts Set of Contacted %"]]) if "Appts Set of Contacted %" in idx else 0
    appt_set_pct = num(row[idx["Appts Set %"]]) if "Appts Set %" in idx else None

    # Percentages alone cannot be aggregated, so derive the underlying counts.
    contacted = int(round(contact_pct * good)) if contact_pct else 0
    if appt_set_pct is not None:
        appts_set = int(round(appt_set_pct * good))
    else:
        appts_set = int(round(set_of_contacted * contacted))

    return {
        "goodLeads": good,
        "sold": num(row[idx["Sold in Time Frame"]]) if "Sold in Time Frame" in idx else 0,
        "apptsShown": num(row[idx["Appts Shown"]]) if "Appts Shown" in idx else 0,
        "contactPct": contact_pct,
        "apptSetOfContactedPct": set_of_contacted,
        "apptSetPct": appt_set_pct,
        "contacted": contacted,
        "apptsSet": appts_set,
    }


def reconcile_total(total, by_lead_type):
    """Rebase the TOTAL row's contacted / appts-set counts onto internet good leads.

    The export copies the internet contact rate onto the TOTAL row, so multiplying it
    by all-lead-type good leads invents contacts that were never reported (509 leads x
    the 60.06% internet rate = 306 "contacts" when only 206 internet leads were
    actually contacted). Engagement is an internet measure throughout this dashboard,
    so the store total carries the internet counts and says so.
    """
    if not total:
        return total
    internet = None
    for node in by_lead_type:
        if (node.get("leadType") or "").strip().lower() == "internet":
            internet = node.get("metrics")
            break
    if not internet:
        total["contacted"] = 0
        total["apptsSet"] = 0
        total["contactScope"] = "none"
        return total
    total["contacted"] = internet.get("contacted", 0)
    total["apptsSet"] = internet.get("apptsSet", 0)
    total["contactScope"] = "internet"
    return total


def parse_kpi(rows, idx):
    """Hierarchical KPI report -> (total, byLeadType). Depth varies by export."""
    total = None
    by_lead_type = []
    cur_lt = None
    cur_inv = None
    has_dealer = "Dealer" in idx

    for row in rows:
        dealer = cell(row, idx, "Dealer")
        lt = cell(row, idx, "Lead Type")
        inv = cell(row, idx, "Inventory Type")
        make = cell(row, idx, "Vehicle Make")

        if dealer == "TOTAL" or lt == "TOTAL":
            total = metrics_from(row, idx)
            continue
        # dealer-level subtotal row (Dealer summary exports)
        if has_dealer and dealer and not lt and not inv and not make:
            if total is None:
                total = metrics_from(row, idx)
            continue
        if lt and not inv and not make:
            cur_lt = {"leadType": lt, "metrics": metrics_from(row, idx), "byInventory": []}
            by_lead_type.append(cur_lt)
            cur_inv = None
        elif lt and inv and not make:
            if cur_lt is None or cur_lt["leadType"] != lt:
                cur_lt = {"leadType": lt, "metrics": None, "byInventory": []}
                by_lead_type.append(cur_lt)
            cur_inv = {"inventoryType": inv, "metrics": metrics_from(row, idx), "byMake": []}
            cur_lt["byInventory"].append(cur_inv)
        elif lt and inv and make and cur_inv is not None:
            cur_inv["byMake"].append({"make": make, "metrics": metrics_from(row, idx)})

    return reconcile_total(total, by_lead_type), by_lead_type


def parse_sales(rows, idx):
    reps, totals = [], None
    for row in rows:
        user = cell(row, idx, "User")
        if not user:
            continue
        if user != "TOTAL" and HOUSE_ACCOUNT.search(user):
            continue  # CRM house accounts are not people
        rec = {
            "name": user,
            "goodLeads": num(row[idx["Good Leads"]]) if "Good Leads" in idx else 0,
            "sold": num(row[idx["Sold in Time Frame"]]) if "Sold in Time Frame" in idx else 0,
            "apptsScheduled": num(row[idx["Appts Scheduled"]]) if "Appts Scheduled" in idx else 0,
            "apptsShown": num(row[idx["Appts Shown"]]) if "Appts Shown" in idx else 0,
            "shownPct": num(row[idx["Appts Shown %"]]) if "Appts Shown %" in idx else 0,
            "calls": num(row[idx["Calls Out"]]) if "Calls Out" in idx else 0,
            "emails": num(row[idx["Emails Out"]]) if "Emails Out" in idx else 0,
            "texts": num(row[idx["Texts Out"]]) if "Texts Out" in idx else 0,
        }
        if user == "TOTAL":
            totals = rec
        else:
            reps.append(rec)
    return reps, totals


def parse_workbook(path):
    """-> snapshot dict, or raises with a reason."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Report" not in wb.sheetnames:
            raise ValueError("no Report sheet")
        filters = read_filters(wb)
        dealer = filters.get("Dealers", "").strip()
        if not dealer or "," in dealer:
            raise ValueError("Filters has no single dealer (%r)" % dealer)
        level = filters.get("Summary Level 1", "").strip()
        if not level:
            raise ValueError("Filters has no Summary Level 1")

        begin = parse_dt(filters.get("Date Range Begin"))
        end = parse_dt(filters.get("Date Range End"))
        run = parse_dt(filters.get("Run Date"))
        if not run:
            raise ValueError("Filters has no usable Run Date")

        date_range = filters.get("Date Range", "").strip()
        if date_range == "Previous Month MTD":
            period = "prior"
        elif date_range == "Current Month":
            period = "current"
        else:
            # Custom Date Range: current if it lands in the run date's own month.
            period = "current" if (begin or "")[:7] == (run or "")[:7] else "prior"

        rows, idx = read_report(wb)
        snap = {
            "storeId": slug(dealer),
            "storeName": dealer,
            "kind": "sales" if level == "User" else "kpi",
            "period": period,
            "dateRange": date_range,
            "begin": begin,
            "end": end,
            "runDate": run,
            "source": os.path.basename(path),
            "rowCount": len(rows),
        }
        if snap["kind"] == "sales":
            reps, totals = parse_sales(rows, idx)
            if not reps and totals is None:
                raise ValueError("sales report had no usable user rows")
            snap["reps"] = reps
            snap["repTotals"] = totals
            snap["total"] = None
            snap["byLeadType"] = []
        else:
            total, by_lt = parse_kpi(rows, idx)
            if total is None and not by_lt:
                raise ValueError("KPI report had no usable rows")
            if total is None:
                # No TOTAL row: sum the lead-type subtotals rather than silently
                # shipping a snapshot with a missing store total.
                total = {k: 0 for k in ("goodLeads", "sold", "apptsShown", "contacted", "apptsSet")}
                for b in by_lt:
                    for k in list(total):
                        total[k] += (b["metrics"] or {}).get(k, 0)
                total["contactPct"] = (total["contacted"] / total["goodLeads"]) if total["goodLeads"] else 0
                total["apptSetOfContactedPct"] = (total["apptsSet"] / total["contacted"]) if total["contacted"] else 0
                total["apptSetPct"] = None
                snap["totalDerived"] = True
            snap["total"] = total
            snap["byLeadType"] = by_lt
            snap["reps"] = None
            snap["repTotals"] = None
        return snap
    finally:
        wb.close()


def parse_matador(path, log):
    out = []
    store = slug(re.sub(r"(?i)^matador\s+mtd\s+stats\s+", "", os.path.basename(path)[:-4]).strip())
    try:
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for r in csv.DictReader(fh):
                out.append({
                    "storeId": store,
                    "name": (r.get("User") or "").strip(),
                    "role": (r.get("User Role") or "").strip(),
                    "lastActivity": (r.get("Last Activity") or "").strip()[:10],
                    "apptsCreated": int(r.get("Appointments Created") or 0),
                    "videosSent": int(r.get("Videos Sent") or 0),
                    "messagesSent": int(r.get("Sent Messages") or 0),
                    "reviewInvites": int(r.get("Review Invites Sent") or 0),
                    "clientsMessaged": int(r.get("Clients Messaged") or 0),
                    "assignedClients": int(r.get("Assigned Clients") or 0),
                })
    except Exception as exc:  # noqa: BLE001
        log.append("SKIP matador %s: %s" % (os.path.basename(path), exc))
    return out


# --------------------------------------------------------------------------- #
# integrations — described, never invented; coverage is computed from the data
# --------------------------------------------------------------------------- #

INTEGRATIONS = [
    {"name": "VinSolutions", "type": "CRM", "api": "requested", "scheduledEmail": True,
     "note": "Enterprise Performance exports arrive daily by email from reportscheduler@motosnap.com. API access requested."},
    {"name": "Tekion", "type": "CRM", "api": "requested", "scheduledEmail": True,
     "note": "Has an API — access inquiry in progress. Reports can be scheduled by email."},
    {"name": "DriveCentric", "type": "CRM", "api": "unknown", "scheduledEmail": True,
     "note": "API availability unconfirmed. Reports can be scheduled."},
    {"name": "Momentum", "type": "CRM", "api": "requested", "scheduledEmail": True,
     "note": "Has an API — asking about access. Reports can be scheduled."},
    {"name": "Matador", "type": "AI messaging", "api": "unknown", "scheduledEmail": False,
     "note": "Texts and videos sent. Activity is exported by hand from the Users tab; no scheduled email today."},
    {"name": "Covideo", "type": "Video outreach", "api": "unknown", "scheduledEmail": True,
     "note": "Activity reporting can be scheduled by email."},
]


def main(argv):
    src_dirs = [os.path.abspath(p) for p in argv[1:]] or [DEFAULT_SRC]
    log = []
    os.makedirs(CACHE, exist_ok=True)

    xlsx_paths, csv_paths = discover(src_dirs, log)
    print("sources: %s" % ", ".join(src_dirs))
    for line in log:
        print("  %s" % line)
    print("workbooks found: %d" % len(xlsx_paths))

    snapshots, skipped = [], []
    excluded = defaultdict(int)
    for path in sorted(xlsx_paths):
        try:
            snap = parse_workbook(path)
        except Exception as exc:  # noqa: BLE001 - one bad file must never abort
            skipped.append((os.path.basename(path), str(exc)))
            continue
        if snap["storeId"] in EXCLUDE_STORES:
            excluded[snap["storeName"]] += 1
            continue
        snapshots.append(snap)

    # de-duplicate identical sends; keep the richer copy
    best = {}
    dupes = 0
    for snap in snapshots:
        key = (snap["storeId"], snap["kind"], snap["period"], snap["begin"], snap["end"], snap["runDate"])
        prev = best.get(key)
        if prev is None:
            best[key] = snap
        else:
            dupes += 1
            if snap["rowCount"] > prev["rowCount"]:
                best[key] = snap
    kept = sorted(best.values(), key=lambda s: (s["storeId"], s["kind"], s["period"], s["runDate"] or ""))

    matador = []
    for path in sorted(csv_paths):
        matador.extend(parse_matador(path, log))
    # drop activity belonging to an excluded store rather than leaving it orphaned
    matador = [m for m in matador if m["storeId"] not in EXCLUDE_STORES]

    # stores and coverage are derived from what actually parsed
    store_names, coverage = {}, defaultdict(lambda: {"runDates": set(), "months": set(), "kinds": set()})
    for snap in kept:
        store_names[snap["storeId"]] = snap["storeName"]
        cov = coverage[snap["storeId"]]
        cov["runDates"].add(snap["runDate"])
        if snap["begin"]:
            cov["months"].add(snap["begin"][:7])
        cov["kinds"].add(snap["kind"])

    matador_stores = {m["storeId"] for m in matador}
    stores = []
    for sid in sorted(store_names):
        stores.append({
            "id": sid,
            "name": store_names[sid],
            "crm": "VinSolutions",
            "tools": ["Matador"] if sid in matador_stores else [],
            "location": None,
        })

    data = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d"),
        "stores": stores,
        "snapshots": kept,
        "matador": matador,
        "integrations": INTEGRATIONS,
        "coverage": {
            sid: {
                "firstRun": min(c["runDates"]) if c["runDates"] else None,
                "lastRun": max(c["runDates"]) if c["runDates"] else None,
                "runDates": sorted(c["runDates"]),
                "months": sorted(c["months"]),
                "kinds": sorted(c["kinds"]),
            }
            for sid, c in coverage.items()
        },
        "ingest": {
            "workbooksFound": len(xlsx_paths),
            "parsed": len(snapshots),
            "unique": len(kept),
            "duplicatesDropped": dupes,
            "skipped": [{"file": f, "reason": r} for f, r in skipped],
        },
    }

    with open(OUT, "w") as fh:
        json.dump(data, fh, indent=1)

    print("parsed %d, unique %d (%d duplicate sends dropped), skipped %d"
          % (len(snapshots), len(kept), dupes, len(skipped)))
    for name, n in sorted(excluded.items()):
        print("  EXCLUDED %s: %d snapshots (listed in EXCLUDE_STORES)" % (name, n))
    for fname, reason in skipped:
        print("  SKIP %s: %s" % (fname, reason))
    print("\nstore coverage")
    for sid in sorted(coverage):
        c = data["coverage"][sid]
        per_kind = defaultdict(int)
        for snap in kept:
            if snap["storeId"] == sid:
                per_kind["%s/%s" % (snap["kind"], snap["period"])] += 1
        detail = ", ".join("%s=%d" % (k, v) for k, v in sorted(per_kind.items()))
        print("  %-20s %s -> %s  (%d run dates)  %s"
              % (store_names[sid], c["firstRun"], c["lastRun"], len(c["runDates"]), detail))
    if matador:
        print("matador rows: %d" % len(matador))
    print("\nwrote %s" % OUT)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
